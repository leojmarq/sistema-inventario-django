import os
import shutil
from datetime import datetime, date
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models
from django.http import HttpResponse, FileResponse
from django.conf import settings
from django.template.loader import get_template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import Producto, Categoria, MovimientoInventario, LogSistema, Proveedor, Cliente, Venta, DetalleVenta
from .decorators import requiere_permiso
from .utils import registrar_log
from .forms import ProductoForm, MovimientoForm
from .categoria_form import CategoriaForm
from .usuario_forms import UsuarioForm, PerfilForm
from .negocio_forms import ProveedorForm, ClienteForm, VentaForm, DetalleVentaForm

@login_required
def inicio(request):
    from .models import Perfil
    
    # Crear perfil automáticamente si no existe
    try:
        perfil = request.user.perfil
    except:
        perfil, created = Perfil.objects.get_or_create(
            user=request.user,
            defaults={'rol': 'admin'}
        )
    
    # Notificaciones de stock bajo
    productos_stock_bajo = Producto.objects.filter(stock__lte=models.F('stock_minimo')).count()
    if productos_stock_bajo > 0:
        messages.warning(request, f'¡Atención! Hay {productos_stock_bajo} producto(s) con stock bajo')
    
    return render(request, 'inventario/inicio.html', {'perfil': perfil})

@login_required
def panel_admin(request):
    from .models import Perfil
    from django.db.models import Sum
    
    # Crear perfil automáticamente si no existe
    try:
        perfil = request.user.perfil
    except:
        perfil, created = Perfil.objects.get_or_create(
            user=request.user,
            defaults={'rol': 'admin'}
        )
        if created:
            messages.success(request, f'Perfil creado automáticamente como {perfil.get_rol_display()}')
    
    if not perfil.puede_ver_reportes():
        messages.error(request, 'No tienes permisos para acceder al panel de administración')
        return redirect('inicio')
    
    # Estadísticas básicas
    total_productos = Producto.objects.count()
    productos_stock_bajo = Producto.objects.filter(stock__lte=models.F('stock_minimo')).count()
    total_categorias = Categoria.objects.count()
    total_movimientos = MovimientoInventario.objects.count()
    
    # Nuevas estadísticas
    total_proveedores = Proveedor.objects.filter(activo=True).count()
    total_clientes = Cliente.objects.filter(activo=True).count()
    ventas_mes = Venta.objects.filter(
        estado='completada',
        fecha__month=datetime.now().month
    ).count()
    ingresos_mes = Venta.objects.filter(
        estado='completada',
        fecha__month=datetime.now().month
    ).aggregate(total=Sum('total'))['total'] or 0
    
    # Productos con stock bajo
    productos_criticos = Producto.objects.filter(stock__lte=models.F('stock_minimo'))[:5]
    
    # Últimos movimientos
    ultimos_movimientos = MovimientoInventario.objects.select_related('producto').order_by('-fecha')[:10]
    
    # Últimas ventas
    ultimas_ventas = Venta.objects.filter(estado='completada').select_related('cliente').order_by('-fecha')[:5]
    
    context = {
        'perfil': perfil,
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'total_categorias': total_categorias,
        'total_movimientos': total_movimientos,
        'total_proveedores': total_proveedores,
        'total_clientes': total_clientes,
        'ventas_mes': ventas_mes,
        'ingresos_mes': ingresos_mes,
        'productos_criticos': productos_criticos,
        'ultimos_movimientos': ultimos_movimientos,
        'ultimas_ventas': ultimas_ventas,
    }
    
    return render(request, 'inventario/panel_admin.html', context)

@login_required
def respaldar_bd(request):
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para respaldar la base de datos')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    try:
        # Crear directorio de respaldos si no existe
        backup_dir = os.path.join(settings.BASE_DIR, 'respaldos')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Nombre del archivo de respaldo con fecha y hora
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'respaldo_{timestamp}.sqlite3'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copiar la base de datos actual
        db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
        shutil.copy2(db_path, backup_path)
        
        messages.success(request, f'Respaldo creado exitosamente: {backup_filename}')
        
        # Registrar log
        registrar_log(request.user, 'respaldo', f'Respaldo de BD creado: {backup_filename}', request)
        
        # Descargar el archivo
        response = FileResponse(
            open(backup_path, 'rb'),
            as_attachment=True,
            filename=backup_filename
        )
        return response
        
    except Exception as e:
        messages.error(request, f'Error al crear respaldo: {str(e)}')
        return redirect('panel_admin')

@login_required
def restaurar_bd(request):
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para restaurar la base de datos')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    if request.method == 'POST':
        if 'backup_file' in request.FILES:
            backup_file = request.FILES['backup_file']
            
            try:
                # Crear respaldo de seguridad antes de restaurar
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                safety_backup = f'respaldo_seguridad_{timestamp}.sqlite3'
                db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
                safety_path = os.path.join(settings.BASE_DIR, 'respaldos', safety_backup)
                
                os.makedirs(os.path.dirname(safety_path), exist_ok=True)
                shutil.copy2(db_path, safety_path)
                
                # Restaurar la base de datos
                with open(db_path, 'wb+') as destination:
                    for chunk in backup_file.chunks():
                        destination.write(chunk)
                
                messages.success(request, 'Base de datos restaurada exitosamente')
                messages.info(request, f'Respaldo de seguridad creado: {safety_backup}')
                
                # Registrar log
                registrar_log(request.user, 'restaurar', f'BD restaurada desde: {backup_file.name}', request)
                return redirect('panel_admin')
                
            except Exception as e:
                messages.error(request, f'Error al restaurar: {str(e)}')
        else:
            messages.error(request, 'No se seleccionó ningún archivo')
    
    return render(request, 'inventario/restaurar_bd.html')

@login_required
def logs_sistema(request):
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para ver los logs')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    usuario_filtro = request.GET.get('usuario')
    accion_filtro = request.GET.get('accion')
    
    logs = LogSistema.objects.all()
    
    if fecha_desde:
        logs = logs.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        logs = logs.filter(fecha__date__lte=fecha_hasta)
    if usuario_filtro:
        logs = logs.filter(usuario__username__icontains=usuario_filtro)
    if accion_filtro:
        logs = logs.filter(accion=accion_filtro)
    
    logs = logs[:100]  # Limitar a 100 registros
    
    context = {
        'logs': logs,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'usuario_filtro': usuario_filtro,
        'accion_filtro': accion_filtro,
        'acciones': LogSistema.ACCIONES,
    }
    
    return render(request, 'inventario/logs_sistema.html', context)

@login_required
def reportes(request):
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para ver reportes')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    # Obtener categorías para el filtro
    categorias = Categoria.objects.all()
    
    return render(request, 'inventario/reportes.html', {'categorias': categorias})

@login_required
def reporte_inventario_pdf(request):
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para generar reportes')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    # Filtros
    categoria_filtro = request.GET.get('categoria')
    stock_bajo = request.GET.get('stock_bajo')
    
    productos = Producto.objects.all()
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    if stock_bajo:
        productos = productos.filter(stock__lte=models.F('stock_minimo'))
    
    # Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Membrete
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    elements.append(Paragraph("SISTEMA DE INVENTARIO", title_style))
    elements.append(Paragraph("Reporte de Inventario", styles['Heading2']))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de productos
    data = [['Código', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Estado']]
    
    for producto in productos:
        estado = 'Stock Bajo' if producto.necesita_restock else 'OK'
        data.append([
            producto.codigo,
            producto.nombre,
            str(producto.categoria),
            f'${producto.precio}',
            str(producto.stock),
            estado
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Registrar log
    registrar_log(request.user, 'crear', 'Reporte de inventario generado en PDF', request)
    
    doc.build(elements)
    return response

@login_required
def exportar_productos_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para exportar')
            return redirect('lista_productos')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('lista_productos')
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Stock Mínimo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    productos = Producto.objects.all().select_related('categoria')
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.codigo)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.categoria.nombre)
        ws.cell(row=row, column=4, value=float(producto.precio))
        ws.cell(row=row, column=5, value=producto.stock)
        ws.cell(row=row, column=6, value=producto.stock_minimo)
        ws.cell(row=row, column=7, value='Stock Bajo' if producto.necesita_restock else 'OK')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    
    registrar_log(request.user, 'crear', 'Productos exportados a Excel', request)
    
    return response

@requiere_permiso('crear')
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            registrar_log(request.user, 'crear', f'Categoría creada: {categoria.nombre}', request)
            messages.success(request, 'Categoría creada exitosamente')
            return redirect('lista_categorias')
    else:
        form = CategoriaForm()
    
    return render(request, 'inventario/crear_categoria.html', {'form': form})

@login_required
def lista_categorias(request):
    categorias = Categoria.objects.all()
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    
    return render(request, 'inventario/lista_categorias.html', {
        'categorias': categorias,
        'perfil': perfil
    })

@login_required
def lista_usuarios(request):
    from django.contrib.auth.models import User
    from .models import Perfil
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para ver usuarios')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    usuarios = User.objects.all().select_related('perfil')
    
    return render(request, 'inventario/lista_usuarios.html', {
        'usuarios': usuarios,
        'perfil': perfil
    })

@login_required
def editar_usuario(request, pk):
    from django.contrib.auth.models import User
    from .models import Perfil
    
    try:
        perfil_actual = request.user.perfil
        if not perfil_actual.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para editar usuarios')
            return redirect('lista_usuarios')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('lista_usuarios')
    
    usuario = get_object_or_404(User, pk=pk)
    
    # Obtener o crear perfil
    perfil, created = Perfil.objects.get_or_create(user=usuario)
    
    if request.method == 'POST':
        usuario_form = UsuarioForm(request.POST, instance=usuario)
        perfil_form = PerfilForm(request.POST, instance=perfil)
        
        if usuario_form.is_valid() and perfil_form.is_valid():
            usuario_form.save()
            perfil_form.save()
            
            registrar_log(request.user, 'editar', f'Usuario editado: {usuario.username}', request)
            messages.success(request, 'Usuario actualizado exitosamente')
            return redirect('lista_usuarios')
    else:
        usuario_form = UsuarioForm(instance=usuario)
        perfil_form = PerfilForm(instance=perfil)
    
    return render(request, 'inventario/editar_usuario.html', {
        'usuario_form': usuario_form,
        'perfil_form': perfil_form,
        'usuario': usuario,
        'perfil': perfil_actual
    })

@login_required
def crear_perfil(request, user_id):
    from django.contrib.auth.models import User
    from .models import Perfil
    
    try:
        perfil_actual = request.user.perfil
        if not perfil_actual.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para crear perfiles')
            return redirect('lista_usuarios')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('lista_usuarios')
    
    usuario = get_object_or_404(User, pk=user_id)
    
    # Verificar que no tenga perfil ya
    if hasattr(usuario, 'perfil'):
        messages.warning(request, 'Este usuario ya tiene un perfil asignado')
        return redirect('editar_usuario', pk=usuario.pk)
    
    if request.method == 'POST':
        perfil_form = PerfilForm(request.POST)
        
        if perfil_form.is_valid():
            perfil = perfil_form.save(commit=False)
            perfil.user = usuario
            perfil.save()
            
            registrar_log(request.user, 'crear', f'Perfil creado para usuario: {usuario.username}', request)
            messages.success(request, 'Perfil creado exitosamente')
            return redirect('lista_usuarios')
    else:
        perfil_form = PerfilForm()
    
    return render(request, 'inventario/crear_perfil.html', {
        'perfil_form': perfil_form,
        'usuario': usuario,
        'perfil': perfil_actual
    })

@login_required
def editar_perfil(request, pk):
    from .models import Perfil
    
    try:
        perfil_actual = request.user.perfil
        if not perfil_actual.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para editar perfiles')
            return redirect('lista_usuarios')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('lista_usuarios')
    
    perfil = get_object_or_404(Perfil, pk=pk)
    
    if request.method == 'POST':
        perfil_form = PerfilForm(request.POST, instance=perfil)
        
        if perfil_form.is_valid():
            perfil_form.save()
            
            registrar_log(request.user, 'editar', f'Perfil editado para usuario: {perfil.user.username}', request)
            messages.success(request, 'Perfil actualizado exitosamente')
            return redirect('lista_usuarios')
    else:
        perfil_form = PerfilForm(instance=perfil)
    
    return render(request, 'inventario/editar_perfil.html', {
        'perfil_form': perfil_form,
        'perfil_editado': perfil,
        'usuario': perfil.user,
        'perfil': perfil_actual
    })

# PROVEEDORES
@login_required
def lista_proveedores(request):
    proveedores = Proveedor.objects.all()
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    return render(request, 'inventario/lista_proveedores.html', {'proveedores': proveedores, 'perfil': perfil})

@requiere_permiso('crear')
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            registrar_log(request.user, 'crear', f'Proveedor creado: {proveedor.nombre}', request)
            messages.success(request, 'Proveedor creado exitosamente')
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'inventario/crear_proveedor.html', {'form': form})

# CLIENTES
@login_required
def lista_clientes(request):
    clientes = Cliente.objects.all()
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    return render(request, 'inventario/lista_clientes.html', {'clientes': clientes, 'perfil': perfil})

@requiere_permiso('crear')
def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            registrar_log(request.user, 'crear', f'Cliente creado: {cliente.nombre}', request)
            messages.success(request, 'Cliente creado exitosamente')
            return redirect('lista_clientes')
    else:
        form = ClienteForm()
    return render(request, 'inventario/crear_cliente.html', {'form': form})

# VENTAS
@login_required
def lista_ventas(request):
    ventas = Venta.objects.all().select_related('cliente', 'usuario')
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    return render(request, 'inventario/lista_ventas.html', {'ventas': ventas, 'perfil': perfil})

@requiere_permiso('crear')
def crear_venta(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            venta = form.save(commit=False)
            venta.usuario = request.user
            venta.save()
            
            registrar_log(request.user, 'crear', f'Venta creada: #{venta.id}', request)
            messages.success(request, 'Venta creada. Ahora agrega productos')
            return redirect('agregar_detalle_venta', venta_id=venta.id)
    else:
        form = VentaForm()
    return render(request, 'inventario/crear_venta.html', {'form': form})

@requiere_permiso('crear')
def agregar_detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    
    if request.method == 'POST':
        form = DetalleVentaForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.venta = venta
            detalle.precio_unitario = detalle.producto.precio
            detalle.save()
            
            # Actualizar stock
            producto = detalle.producto
            producto.stock -= detalle.cantidad
            producto.save()
            
            # Recalcular total
            venta.calcular_total()
            
            messages.success(request, 'Producto agregado a la venta')
            return redirect('agregar_detalle_venta', venta_id=venta.id)
    else:
        form = DetalleVentaForm()
    
    detalles = venta.detalleventa_set.all()
    return render(request, 'inventario/agregar_detalle_venta.html', {
        'form': form,
        'venta': venta,
        'detalles': detalles
    })

@requiere_permiso('editar')
def finalizar_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    venta.estado = 'completada'
    venta.save()
    
    registrar_log(request.user, 'editar', f'Venta finalizada: #{venta.id}', request)
    messages.success(request, 'Venta finalizada exitosamente')
    return redirect('lista_ventas')

@login_required
def reporte_ventas_pdf(request):
    from .models import Perfil
    from datetime import datetime, timedelta
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para generar reportes')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    ventas = Venta.objects.filter(estado='completada')
    
    if fecha_desde:
        ventas = ventas.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha__date__lte=fecha_hasta)
    
    # Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Membrete
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    
    elements.append(Paragraph("SISTEMA DE INVENTARIO", title_style))
    elements.append(Paragraph("Reporte de Ventas", styles['Heading2']))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Estadísticas
    total_ventas = ventas.count()
    total_ingresos = sum(v.total for v in ventas)
    
    elements.append(Paragraph(f"<b>Total de Ventas:</b> {total_ventas}", styles['Normal']))
    elements.append(Paragraph(f"<b>Ingresos Totales:</b> ${total_ingresos:.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de ventas
    data = [['#', 'Cliente', 'Fecha', 'Total']]
    
    for venta in ventas:
        data.append([
            str(venta.id),
            venta.cliente.nombre,
            venta.fecha.strftime('%d/%m/%Y'),
            f'${venta.total}'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    registrar_log(request.user, 'crear', 'Reporte de ventas generado en PDF', request)
    
    doc.build(elements)
    return response

@login_required
def reporte_productos_mas_vendidos(request):
    from .models import Perfil
    from django.db.models import Sum, Count
    
    try:
        perfil = request.user.perfil
        if not perfil.puede_ver_reportes():
            messages.error(request, 'No tienes permisos para ver reportes')
            return redirect('panel_admin')
    except:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('panel_admin')
    
    # Top 10 productos más vendidos
    productos_vendidos = DetalleVenta.objects.filter(
        venta__estado='completada'
    ).values(
        'producto__nombre'
    ).annotate(
        total_vendido=Sum('cantidad'),
        veces_vendido=Count('id')
    ).order_by('-total_vendido')[:10]
    
    return render(request, 'inventario/reporte_productos_vendidos.html', {
        'productos_vendidos': productos_vendidos
    })

@login_required
def lista_productos(request):
    productos = Producto.objects.all()
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    
    # Búsqueda avanzada
    buscar = request.GET.get('buscar')
    categoria_filtro = request.GET.get('categoria')
    stock_filtro = request.GET.get('stock')
    
    if buscar:
        productos = productos.filter(
            models.Q(nombre__icontains=buscar) | 
            models.Q(codigo__icontains=buscar) |
            models.Q(descripcion__icontains=buscar)
        )
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    
    if stock_filtro == 'bajo':
        productos = productos.filter(stock__lte=models.F('stock_minimo'))
    elif stock_filtro == 'alto':
        productos = productos.filter(stock__gt=models.F('stock_minimo'))
    
    categorias = Categoria.objects.all()
    
    return render(request, 'inventario/lista_productos.html', {
        'productos': productos,
        'perfil': perfil,
        'categorias': categorias,
        'buscar': buscar,
        'categoria_filtro': categoria_filtro,
        'stock_filtro': stock_filtro
    })

@login_required
def detalle_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    movimientos = MovimientoInventario.objects.filter(producto=producto).order_by('-fecha')[:10]
    
    try:
        perfil = request.user.perfil
    except:
        perfil = None
    
    return render(request, 'inventario/detalle_producto.html', {
        'producto': producto,
        'movimientos': movimientos,
        'perfil': perfil
    })

@requiere_permiso('crear')
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            
            # Registrar log
            registrar_log(request.user, 'crear', f'Producto creado: {producto.nombre} ({producto.codigo})', request)
            
            messages.success(request, 'Producto creado exitosamente')
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    
    categorias = Categoria.objects.all()
    return render(request, 'inventario/crear_producto.html', {
        'form': form,
        'categorias': categorias
    })

@requiere_permiso('editar')
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            
            # Registrar log
            registrar_log(request.user, 'editar', f'Producto editado: {producto.nombre} ({producto.codigo})', request)
            
            messages.success(request, 'Producto actualizado exitosamente')
            return redirect('detalle_producto', pk=producto.pk)
    else:
        form = ProductoForm(instance=producto)
    
    categorias = Categoria.objects.all()
    return render(request, 'inventario/editar_producto.html', {
        'form': form,
        'producto': producto,
        'categorias': categorias
    })

@requiere_permiso('eliminar')
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        nombre_producto = producto.nombre
        codigo_producto = producto.codigo
        producto.delete()
        
        # Registrar log
        registrar_log(request.user, 'eliminar', f'Producto eliminado: {nombre_producto} ({codigo_producto})', request)
        
        messages.success(request, 'Producto eliminado exitosamente')
        return redirect('lista_productos')
    
    return render(request, 'inventario/eliminar_producto.html', {'producto': producto})

@requiere_permiso('editar')
def registrar_movimiento(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = MovimientoForm(request.POST, producto=producto)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.producto = producto
            movimiento.save()
            
            # Actualizar stock
            if movimiento.tipo == 'entrada':
                producto.stock += movimiento.cantidad
            else:
                producto.stock -= movimiento.cantidad
            
            producto.save()
            
            # Registrar log
            registrar_log(request.user, 'editar', f'Movimiento registrado: {movimiento.tipo} de {movimiento.cantidad} unidades - {producto.nombre}', request)
            
            messages.success(request, f'Movimiento de {movimiento.tipo} registrado')
            return redirect('detalle_producto', pk=producto.pk)
    else:
        form = MovimientoForm()
    
    return render(request, 'inventario/registrar_movimiento.html', {
        'form': form,
        'producto': producto
    })